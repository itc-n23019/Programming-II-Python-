import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

for row_num in range(2, sheet.max_row): # 先頭行をスキップ ❶
    produce_name = sheet.cell(row=row_num, column=1).value  # ❷
    if produce_name in PRICE_UPDATES:  # ❸
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save('updatedProduceSales.xlsx')  # ❹

