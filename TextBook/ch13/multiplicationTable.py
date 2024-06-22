import openpyxl
from openpyxl.styles import Font
import sys

if len(sys.argv) != 2:
    sys.exit('Usage: python multiplicationTable.py N')

n = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.active
bold_font = Font(size=12, bold=True)

for i in range(1, n + 1):
    sheet.cell(row=i + 1, column=1, value=i).font = bold_font
    sheet.cell(row=1, column=i + 1, value=i).font = bold_font
    for j in range(1, n + 1):
        sheet.cell(row=i + 1, column=j + 1, value=i * j)

wb.save('multiplicationtable.xlsx')

